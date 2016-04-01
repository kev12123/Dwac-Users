from  openpyxl import *
import string
class Excel:

    def __init__(self):
        self.wb = Workbook()


    def transform_to_excel_spreadSheet(self,data,spreadsheet_path):
        data_set = [row for row in data ]
        sheet_columns = list(string.ascii_uppercase)
        column_names = []
        #GET COLUMNS NAMES
        for data in data_set:
            for column,value in data.items():
                column_names.append(column)
            break
        #INSERT COLUMN NAMES INTO THE SPREADSHEET
        ws = self.wb.active
        for column in range(len(column_names)):
            ws[sheet_columns[column]+'1']=column_names[column]

        number_of_rows = len([row[0] for row in data_set]) #GET NUMBER OF ROWS IN SPREADSHEET

        #INSERT DATA INTO SPREADSHEET

        for column in range(len(column_names)-1):
            selected_column = sheet_columns[column]
            for row in range(number_of_rows):
                ws[selected_column+str(row+2)] =data_set[row][column]

        # FIX CELL STYLING FROM SPREADSHEET
        columns_widths = ['' for x in column_names]
        for column in range(len(column_names)):
            max_length = 0
            for row in ws.iter_rows(str(sheet_columns[0]+'1')+":"+str(sheet_columns[len(column_names)-1]+str(len(column_names)-1))):
                for cell in row:
                    if len(str(cell)) > max_length:
                        max_length = len(str(cell))
                    columns_widths[column] = max_length

        for i , columns_width in enumerate(columns_widths):
            ws.column_dimensions[sheet_columns[i]].width = columns_width




        self.wb.save(spreadsheet_path)

