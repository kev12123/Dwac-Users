from sqlalchemy import *
import unicodedata
import urllib
from openpyxl import *
from openpyxl.styles import colors
from openpyxl.styles import Font, Color,Fill,PatternFill,Border
import string
from Excel import *

def load_connection( driver ='SQL Server Native Client 10.0',server='astsqlent02dev',database = 'TeamWorkShop'):
        #Function to load sessiona and MetaData
        quoted = urllib.quote_plus('DRIVER={0};Server={1};Database={2};Trusted_Connection=yes;'.format(driver,server,database))
        connection =create_engine('mssql+pyodbc:///?odbc_connect={}'.format(quoted))
        return connection

def equals(a,b):
    if str(a).strip().lower() == str(b).strip().lower():
        return True
    return False

conn = load_connection()
query = conn.execute("SELECT * FROM ceridian")
ceridian_data = [data for data in query]
ceridian_wb = Excel()
ceridian_wb.transform_to_excel_spreadSheet(query,'C:\Users\kgiraldo\Desktop\practice_table.xlsx')


# output_sql_table_to_excel(query,'C:\Users\kgiraldo\Desktop\ceridian_table.xlsx')
query = conn.execute('SELECT * FROM employees')
dwack_data = [data for data in query]
# output_sql_table_to_excel('query,'C:\Users\kgiraldo\Desktop\dwack_users.xlsx')
book = load_workbook('C:\Users\kgiraldo\Desktop\dwack_users.xlsx')
ws=book.get_sheet_by_name('Sheet')

#EMPLOYEES WITH USER IDS IN THE CERIDIAN  TABLE
for i in ceridian_data:
    user_ids = i[13]
    for row in ws.iter_rows('C2:C99'):
        for cell in row:
            if str(cell.internal_value).strip().lower() == str(user_ids).strip().lower():
                ws['E' + str(cell.row)]=str(i[5])
                ws['F' + str(cell.row)]=str(i[10])
                ws['G' + str(cell.row)]=str(i[6])

#EMPLOYEES WITHOUT USERID IN THE CERIDIAN TABLE
#THE NAMING COVENTION FOR THE USERID IS INITIAL OF FIRST NAME AND LAST NAME
#USING THIS PATTERN I WILL FIND CREATE THE USER ID FROM THE CERIDDIAN LIST COLUMN
#AND FIND THE REMAINING MATCHES
for i in ceridian_data:
    if i[13] is None:
        last_name,first_name =i[5].split(',')
        user_id = first_name[:2]+last_name.strip()
        for row in ws.iter_rows('C2:C99'):
            for cell in row:
                if equals(cell.value,user_id.strip):
                     ws['E' + str(cell.row)]=str(i[5])
                     ws['F' + str(cell.row)]=str(i[10])
                     ws['G' + str(cell.row)]=str(i[6])

#STYLE SPREADSHEET FORMATTTING

num_of_columns = 1
for data in dwack_data:
    for column,value in data.items():
        num_of_columns +=1
    break

print num_of_columns

sheet_columns = list(string.ascii_uppercase)

columns_widths = ['' for x in range(num_of_columns)]
for column in range(num_of_columns):
        max_length =0
        for row in ws.iter_rows(str(sheet_columns[0]+'1')+":"+str(sheet_columns[num_of_columns-1]+str(num_of_columns-1))):
            for cell in row:
                if len(str(cell)) > max_length:
                    max_length = len(str(cell))
                columns_widths[column]= max_length


#FORMATTING SPREADSHEET STYLE

for column_width in columns_widths:
    for i in range(num_of_columns):
        ws.column_dimensions[sheet_columns[i]].width = 33

redFill = PatternFill(start_color=colors.BLACK,
                   end_color= colors.BLACK,
                   fill_type='solid')
for i in range(8):
    ws[sheet_columns[i]+'1'].fill = redFill
    ws[sheet_columns[i]+'1'].font =Font(color=colors.WHITE)



book.save('C:\Users\kgiraldo\Desktop\dwack_usdddd.xlsx')













